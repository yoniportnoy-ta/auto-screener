"""One-shot importer for the legacy Apps Script auto-screener data.

What we migrate:
  * Per-class feedback tabs from the "Screener Feedback" Google Sheet, accepted
    in EITHER format:
      - a directory of *.csv exports (one CSV per class — Google Sheets only
        exports the active tab as CSV)
      - a single *.xlsx export of the whole workbook ("File → Download → Excel"),
        which contains all tabs at once — recommended.
  * Position-class assignments from Apps Script properties (a JSON dump
    keyed by position_uid).

Idempotent: re-running won't duplicate feedback (we de-dupe by timestamp +
candidate_uid), and re-importing position-class assignments simply UPSERTs.

Usage from the Render web service Shell:

    # 1. Upload exports + class-map JSON to /tmp.

    # 2. Run the imports (any of the three feedback forms):
    python -m app.migrate_apps_script feedback /tmp/migration/feedback           # folder of CSVs
    python -m app.migrate_apps_script feedback /tmp/migration/screener.xlsx      # single workbook
    python -m app.migrate_apps_script feedback /tmp/migration/Backend.csv        # one tab as CSV
    python -m app.migrate_apps_script classes  /tmp/migration/position_classes.json
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import sys
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from .db import db_session
from .logging_config import configure_logging
from .models import Feedback, PositionClass
from .position_classes import DEFAULT_CLASSES, list_all_classes

log = logging.getLogger(__name__)

# Apps Script feedback sheet column order — see Code.gs FEEDBACK_COLS_:
#   Timestamp | RecruiterEmail | PositionUID | PositionName | CandidateUID
#   | CandidateName | AIRating | Verdict | Note
EXPECTED_HEADERS = (
    "Timestamp", "RecruiterEmail", "PositionUID", "PositionName",
    "CandidateUID", "CandidateName", "AIRating", "Verdict", "Note",
)


@dataclass
class ImportSummary:
    files_processed: int = 0
    rows_seen: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    skipped_reasons: dict[str, int] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    def bump_skip(self, reason: str) -> None:
        self.skipped_reasons[reason] = self.skipped_reasons.get(reason, 0) + 1
        self.rows_skipped += 1


# ─── Feedback importer (CSV or XLSX) ─────────────────────────────────────────
# Tabs in the Google Sheet that aren't class feedback (skip them in XLSX import).
NON_CLASS_TABS = {"_LearnedRubrics", "_DebugScoring", "_DebugScan", "_Notes", "Sheet1"}


def import_feedback(path: Path) -> ImportSummary:
    """Dispatch on the input path: file vs. directory, CSV vs. XLSX."""
    if path.is_dir():
        return import_feedback_csvs(path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return import_feedback_xlsx(path)
    if suffix == ".csv":
        return import_feedback_csvs(path.parent, only=[path])
    summary = ImportSummary()
    summary.errors.append(f"unrecognised input: {path} (expected directory, .xlsx, or .csv)")
    return summary


def import_feedback_csvs(directory: Path, *, only: list[Path] | None = None) -> ImportSummary:
    """Read every *.csv in `directory` (or just `only` if provided). Each filename
    should match the class tab name (e.g. "Backend.csv", "Product Management.csv").
    """
    summary = ImportSummary()
    if not directory.is_dir() and only is None:
        summary.errors.append(f"not a directory: {directory}")
        return summary

    paths = list(only) if only else sorted(directory.glob("*.csv"))
    for path in paths:
        summary.files_processed += 1
        class_id, class_name = _resolve_class_from_label(path.stem.strip())

        try:
            with path.open("r", newline="", encoding="utf-8-sig") as fh:
                rows = list(csv.reader(fh))
        except Exception as exc:  # noqa: BLE001
            summary.errors.append(f"{path.name}: read failed: {exc}")
            continue

        _import_rows_for_tab(
            tab_label=path.name,
            rows=rows,
            class_id=class_id,
            class_name=class_name,
            summary=summary,
        )

    return summary


def import_feedback_xlsx(xlsx_path: Path) -> ImportSummary:
    """Read every tab in an .xlsx workbook and import each as a class's feedback."""
    summary = ImportSummary()
    if not xlsx_path.is_file():
        summary.errors.append(f"not a file: {xlsx_path}")
        return summary
    try:
        from openpyxl import load_workbook
    except ImportError:
        summary.errors.append("openpyxl not installed (pip install openpyxl)")
        return summary

    try:
        wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        summary.errors.append(f"openpyxl load: {exc}")
        return summary

    for sheet in wb.worksheets:
        tab_name = (sheet.title or "").strip()
        if not tab_name or tab_name in NON_CLASS_TABS or tab_name.startswith("_"):
            log.info("xlsx: skipping non-class tab %r", tab_name)
            continue
        summary.files_processed += 1

        rows: list[list[str]] = []
        for raw_row in sheet.iter_rows(values_only=True):
            cells = [(_cell_to_str(v)) for v in raw_row]
            # Drop trailing empty cells the writer pads.
            while cells and cells[-1] == "":
                cells.pop()
            if not cells:
                continue
            rows.append(cells)
        if not rows:
            continue

        class_id, class_name = _resolve_class_from_label(tab_name)
        _import_rows_for_tab(
            tab_label=f"{xlsx_path.name} :: {tab_name}",
            rows=rows,
            class_id=class_id,
            class_name=class_name,
            summary=summary,
        )

    return summary


def _resolve_class_from_label(label: str) -> tuple[str, str]:
    """Map a tab name / CSV filename stem to a (class_id, class_name)."""
    name_to_id = {c["name"]: c["id"] for c in list_all_classes()}
    id_to_id = {c["id"]: c["id"] for c in list_all_classes()}
    label_clean = label.strip()
    class_id = name_to_id.get(label_clean) or id_to_id.get(label_clean.lower())
    if class_id is None:
        log.warning("import_feedback: class %r not in catalogue; using class_id='general'", label_clean)
        return ("general", label_clean or "General")
    return (class_id, label_clean)


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _import_rows_for_tab(
    *,
    tab_label: str,
    rows: list[list[str]],
    class_id: str,
    class_name: str,
    summary: ImportSummary,
) -> None:
    if not rows:
        summary.errors.append(f"{tab_label}: empty")
        return
    if rows[0] != list(EXPECTED_HEADERS):
        summary.errors.append(
            f"{tab_label}: header mismatch (expected {EXPECTED_HEADERS}, got {rows[0]})"
        )
        return

    log.info("importing %s as class_id=%s class_name=%s", tab_label, class_id, class_name)
    for raw in rows[1:]:
        summary.rows_seen += 1
        if not raw:
            continue
        row = (list(raw) + [""] * len(EXPECTED_HEADERS))[: len(EXPECTED_HEADERS)]
        ts_raw, recruiter_email, pos_uid, pos_name, cand_uid, cand_name, ai_str, rec_str, note = row
        if not (cand_uid and pos_uid):
            summary.bump_skip("missing candidate or position uid")
            continue

        ai_rating = _coerce_rating(ai_str)
        rec_rating = _coerce_rating(rec_str)
        if rec_rating is None and ai_rating is None:
            summary.bump_skip("no ratings")
            continue

        ts = _parse_iso(ts_raw)
        if _feedback_already_imported(ts, cand_uid, rec_rating):
            summary.bump_skip("duplicate")
            continue

        try:
            with db_session() as ses:
                ses.add(Feedback(
                    timestamp=ts or datetime.now(timezone.utc),
                    recruiter_email=(recruiter_email or "").strip()[:200],
                    class_id=class_id,
                    class_name=class_name,
                    position_uid=pos_uid.strip(),
                    position_name=(pos_name or "").strip(),
                    candidate_uid=cand_uid.strip(),
                    candidate_name=(cand_name or "").strip(),
                    ai_rating=ai_rating,
                    recruiter_rating=rec_rating,
                    note=(note or "").strip()[:2000],
                ))
            summary.rows_imported += 1
        except Exception as exc:  # noqa: BLE001
            summary.errors.append(f"{tab_label}:{summary.rows_seen}: insert failed: {exc}")


# ─── Position-class JSON importer ────────────────────────────────────────────
def import_position_classes(json_path: Path) -> ImportSummary:
    """Read a JSON file shaped like:

        {
          "<position_uid>": {"classId": "backend", "className": "Backend", "level": ""},
          ...
        }

    Apps Script stored these in `SCREENER_POS_CLASS_<uid>` script properties; export
    them via `PropertiesService.getScriptProperties().getProperties()` and post-process
    into this shape. UPSERTs into position_classes.
    """
    summary = ImportSummary()
    if not json_path.is_file():
        summary.errors.append(f"not a file: {json_path}")
        return summary

    try:
        data = json.loads(json_path.read_text())
    except Exception as exc:  # noqa: BLE001
        summary.errors.append(f"json parse: {exc}")
        return summary

    if not isinstance(data, dict):
        summary.errors.append("expected JSON object keyed by position_uid")
        return summary

    catalogue_ids = {c["id"] for c in list_all_classes()}
    name_lookup = {c["id"]: c["name"] for c in list_all_classes()}
    custom_seen: set[str] = set()  # avoid duplicate custom-class inserts in one run

    for position_uid, payload in data.items():
        summary.rows_seen += 1
        if not isinstance(payload, dict):
            summary.bump_skip("payload not an object")
            continue
        class_id = (payload.get("classId") or payload.get("class_id") or "").strip()
        class_name = (
            payload.get("className") or payload.get("class_name") or name_lookup.get(class_id) or ""
        ).strip()
        level = (payload.get("level") or "").strip() or None
        if not class_id or not class_name:
            summary.bump_skip("missing class_id or class_name")
            continue

        # Auto-register classes that aren't in the default catalogue (e.g. legacy
        # values like "it", "nlp", "talent_acquisition" that the recruiter created
        # in the Apps Script version). Otherwise the UI dropdown wouldn't show them.
        if class_id not in catalogue_ids and class_id not in custom_seen:
            _ensure_custom_class(class_id, class_name)
            custom_seen.add(class_id)

        with db_session() as ses:
            stmt = pg_insert(PositionClass).values(
                position_uid=position_uid.strip(),
                class_id=class_id,
                class_name=class_name,
                level=level,
            )
            stmt = stmt.on_conflict_do_update(
                index_elements=[PositionClass.position_uid],
                set_={
                    "class_id": stmt.excluded.class_id,
                    "class_name": stmt.excluded.class_name,
                    "level": stmt.excluded.level,
                },
            )
            ses.execute(stmt)
        summary.rows_imported += 1

    return summary


def _ensure_custom_class(class_id: str, class_name: str) -> None:
    """Insert a custom class row if not already present. Used by the position-class
    importer to register classes that exist in legacy data but not in the catalogue."""
    from .models import CustomPositionClass

    with db_session() as ses:
        existing = ses.scalar(
            select(CustomPositionClass).where(CustomPositionClass.class_id == class_id)
        )
        if existing:
            return
        stmt = pg_insert(CustomPositionClass).values(
            class_id=class_id, class_name=class_name, levels_json=[],
        )
        stmt = stmt.on_conflict_do_nothing(index_elements=[CustomPositionClass.class_id])
        ses.execute(stmt)
    log.info("auto-registered custom class: id=%s name=%s", class_id, class_name)


# ─── Helpers ─────────────────────────────────────────────────────────────────
def _coerce_rating(raw: str | None) -> int | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        n = int(float(s))
        if 1 <= n <= 5:
            return n
    except (ValueError, TypeError):
        pass
    return None


def _parse_iso(raw: str) -> datetime | None:
    s = (raw or "").strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


def _feedback_already_imported(ts: datetime | None, candidate_uid: str, recruiter_rating: int | None) -> bool:
    """Cheap de-dupe: same (timestamp, candidate, recruiter rating) tuple."""
    if ts is None:
        return False
    with db_session() as ses:
        existing = ses.scalar(
            select(Feedback).where(
                Feedback.timestamp == ts,
                Feedback.candidate_uid == candidate_uid,
                Feedback.recruiter_rating == recruiter_rating,
            )
        )
        return existing is not None


# ─── CLI ─────────────────────────────────────────────────────────────────────
def main() -> int:
    configure_logging()
    parser = argparse.ArgumentParser(prog="app.migrate_apps_script")
    sub = parser.add_subparsers(dest="cmd", required=True)
    fb = sub.add_parser("feedback", help="import feedback from a CSV folder, single CSV, or XLSX workbook")
    fb.add_argument("directory", help="path to a feedback directory, .xlsx, or .csv file")
    cls = sub.add_parser("classes", help="import position-class assignments from a JSON dump")
    cls.add_argument("json_path")
    args = parser.parse_args()

    if args.cmd == "feedback":
        summary = import_feedback(Path(args.directory))
    elif args.cmd == "classes":
        summary = import_position_classes(Path(args.json_path))
    else:
        return 2

    log.info(
        "migrate %s done: files=%d seen=%d imported=%d skipped=%d errors=%d",
        args.cmd, summary.files_processed, summary.rows_seen, summary.rows_imported,
        summary.rows_skipped, len(summary.errors),
    )
    if summary.skipped_reasons:
        log.info("skipped breakdown: %s", summary.skipped_reasons)
    if summary.errors:
        for err in summary.errors[:20]:
            log.warning("  err: %s", err)
        if len(summary.errors) > 20:
            log.warning("  … (%d more errors elided)", len(summary.errors) - 20)
    return 0 if not summary.errors else 1


if __name__ == "__main__":
    sys.exit(main())
